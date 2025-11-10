import requests
import json
import pandas as pd
from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import csv

# ---------- 1. Adım: CSV'yi API'den çek ----------

url_api = "https://www.kulturportali.gov.tr/Moduller/GezilecekYerler.aspx/GezilecekYerleriFilitreliGetir"
payload = {
    "sira": 1,
    "sayi": 1,
    "TurKod": "0",
    "TurizmTurKod": "0",
    "ilID": 0,
    "gorsel": False,
    "nearest": False,
    "aramaText": "",
    "etiket": "",
    "HariciEtiket": "",
    "lat": "0",
    "lang": "0"
}

headers = {
    "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:136.0) Gecko/20100101 Firefox/136.0",
    "Accept": "*/*",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate, br",
    "X-Requested-With": "XMLHttpRequest",
    "Content-Type": "application/json; charset=utf-8",
    "Origin": "https://www.kulturportali.gov.tr",
    "Connection": "keep-alive",
    "Referer": "https://www.kulturportali.gov.tr/turkiye/genel/gezilecekyer",
}

resp = requests.post(url_api, json=payload, headers=headers)
resp.raise_for_status()
kayit_data = resp.json()
kayit_entry = json.loads(kayit_data['d'])[0]
kayit_sayisi = kayit_entry["KayitSayisi"]

payload["sayi"] = kayit_sayisi
response = requests.post(url_api, json=payload, headers=headers)
response.raise_for_status()
data = response.json()
entries = json.loads(data['d'])

# CSV oluştur
with open('cultural_places.csv', 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['Name', 'Image', 'Url'])
    for entry in entries:
        writer.writerow([entry["Baslik"], entry["Resim"], entry['Url']])

# ---------- 2. Adım: Her URL için detayları çek ----------

base_url = "https://www.kulturportali.gov.tr"
df = pd.read_csv("cultural_places.csv")

df["Type"] = ""
df["Description"] = ""

for index, row in df.iterrows():
    full_url = base_url + row["Url"]
    try:
        response = requests.get(full_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")

            # Tür'ü çek
            type_span = soup.select_one("div.col-md-6.col-sm-12 label:contains('Tür:') + span")
            if type_span:
                df.at[index, "Type"] = type_span.get_text(strip=True)

            # Açıklama
            description_div = soup.find("div", id="descriptionDiv")
            if description_div:
                # <b> ve <strong> içindeki metinleri escape et
                for bold_tag in description_div.find_all(["b", "strong"]):
                    bold_text = bold_tag.get_text()
                    bold_tag.string = bold_text.replace('"', r'\"')

                # Metni al, paragraf ve br'leri koru
                description_lines = []
                for elem in description_div.descendants:
                    if isinstance(elem, str):
                        description_lines.append(elem)
                    elif elem.name in ["br", "p"]:
                        description_lines.append("\n")
                description_text = "".join(description_lines).strip()
                df.at[index, "Description"] = description_text

    except Exception:
        continue

# ---------- 3. Adım: Excel'e yaz ----------

wb = Workbook()
ws = wb.active
ws.title = "Cultural Places"

ws.append(["Başlık", "Fotoğraf", "Tür", "Açıklama"])
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 100

def px_to_excel_row(px):
    return px * 0.75 / 1.0

for index, row in df.iterrows():
    ws.append([row["Name"], "", row["Type"], row["Description"]])

    full_image_url = base_url + row["Image"]
    try:
        response = requests.get(full_image_url)
        if response.status_code == 200:
            img_data = BytesIO(response.content)
            img = XLImage(img_data)
            img.anchor = f'B{ws.max_row}'
            ws.add_image(img)
            ws.row_dimensions[ws.max_row].height = px_to_excel_row(img.height)
    except Exception:
        continue

wb.save("cultural_places.xlsx")
